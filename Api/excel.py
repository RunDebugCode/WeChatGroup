# -*- coding:UTF-8 -*-
import openpyxl

from Config import settings


def read(path, sheet=0):
    workbook = openpyxl.load_workbook(path)
    worksheet = workbook.worksheets[sheet]

    data = []
    for row in worksheet.iter_rows():
        data.append([cell.value for cell in row])
    workbook.close()
    return data


def write(name, data, save_path):
    workbook = openpyxl.Workbook()
    worksheet = workbook.create_sheet(name)
    title = settings.TITLE
    data1 = title + data

    for i, value_i in enumerate(data1):
        if i != 0:
            data2 = [i] + value_i
            for j, value_j in enumerate(data2):
                cell = worksheet.cell(row=i+1, column=j+1)
                cell.value = value_j
        else:
            for j, value_j in enumerate(value_i):
                cell = worksheet.cell(row=i+1, column=j+1)
                cell.value = value_j
    del workbook['Sheet']
    workbook.save(save_path)


def main():
    pass


if __name__ == '__main__':
    main()
